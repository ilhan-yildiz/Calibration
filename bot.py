import os
import logging
from flask import Flask, jsonify
import threading
import openpyxl
from openpyxl.styles import PatternFill
import requests
from io import BytesIO
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes
from datetime import datetime
import asyncio
import base64
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
EXCEL_URL = os.getenv("EXCEL_URL")
SHEET_NAME = os.getenv("SHEET_NAME", "Tx Detail List")
PORT = int(os.environ.get('PORT', 10000))

# GitHub konfigürasyonu
GITHUB_PAT_TOKEN = os.getenv("GITHUB_PAT_TOKEN")
GITHUB_REPO = os.getenv("GITHUB_REPO", "ilhan-yildiz/Calibration")
GITHUB_FILE_PATH = os.getenv("GITHUB_FILE_PATH", "Yearly_Calibration_Schedule.xlsx")

# Flask uygulaması
flask_app = Flask(__name__)

@flask_app.route('/')
def health_check():
    return jsonify({"status": "ok"}), 200

def run_http_server():
    flask_app.run(host='0.0.0.0', port=PORT, use_reloader=False)

threading.Thread(target=run_http_server, daemon=True).start()
logger.info(f"HTTP sunucusu başlatıldı - Port: {PORT}")

def format_date_to_tr(date_value):
    """Tarihi gg.aa.yyyy formatına çevir (Excel'deki yyyy-aa-gg'den)"""
    if not date_value:
        return None
    
    if isinstance(date_value, str):
        if '-' in date_value:
            parts = date_value.split('-')
            if len(parts) == 3:
                return f"{parts[2]}.{parts[1]}.{parts[0]}"
        elif '.' in date_value:
            if re.match(r'^\d{2}\.\d{2}\.\d{4}$', date_value):
                return date_value
        return None
    
    if hasattr(date_value, 'strftime'):
        return date_value.strftime("%d.%m.%Y")
    
    return None

def validate_date_tr(date_string):
    """gg.aa.yyyy formatını kontrol et ve yyyy-aa-gg formatına çevir"""
    pattern = r'^(\d{2})\.(\d{2})\.(\d{4})$'
    match = re.match(pattern, date_string)
    if match:
        day, month, year = match.groups()
        try:
            datetime(int(year), int(month), int(day))
            return f"{year}-{month}-{day}", True
        except ValueError:
            return None, False
    return None, False

def parse_date_tr(date_string):
    """gg.aa.yyyy formatını datetime objesine çevir"""
    try:
        day, month, year = date_string.split('.')
        return datetime(int(year), int(month), int(day))
    except:
        return None

def get_column_headers(workbook, sheet_name):
    """2. satırdaki kolon başlıklarını al"""
    try:
        sheet = workbook[sheet_name]
        headers = {}
        for col_idx in range(1, sheet.max_column + 1):
            header_value = sheet.cell(2, col_idx).value
            if header_value:
                headers[col_idx] = header_value
        return headers
    except:
        return {}

def search_in_column_c_partial(search_value, workbook, sheet_name):
    """C sütununda kısmi eşleşme arama - Tablo formatında çıktı"""
    try:
        if sheet_name not in workbook.sheetnames:
            return None, f"❌ '{sheet_name}' sayfası bulunamadı!\nMevcut sayfalar: {', '.join(workbook.sheetnames)}"
        
        sheet = workbook[sheet_name]
        headers = get_column_headers(workbook, sheet_name)
        results = []
        search_lower = str(search_value).lower().strip()
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
            if len(row) > 2 and row[2] is not None:
                cell_value = str(row[2]).strip()
                if search_lower in cell_value.lower():
                    table = "```\n"
                    table += "┌────────────┬─────────────────────────────────┐\n"
                    table += "│ Alan       │ Değer                           │\n"
                    table += "├────────────┼─────────────────────────────────┤\n"
                    table += f"│ Satır No   │ {row_idx:<31} │\n"
                    
                    for col_idx, header in headers.items():
                        if col_idx - 1 < len(row) and row[col_idx - 1] is not None:
                            value = str(row[col_idx - 1])
                            if col_idx == 5:
                                formatted = format_date_to_tr(value)
                                value = formatted if formatted else "Belirtilmemiş"
                            if len(value) > 30:
                                value = value[:27] + "..."
                            table += f"├────────────┼─────────────────────────────────┤\n"
                            table += f"│ {header:<10} │ {value:<31} │\n"
                    
                    table += "└────────────┴─────────────────────────────────┘\n"
                    table += "```"
                    results.append(table)
                    
                    if len(results) >= 10:
                        break
        
        if not results:
            return None, None
        return results, None
        
    except Exception as e:
        return None, f"Arama hatası: {str(e)}"

def search_calibration_date(search_value, workbook, sheet_name):
    """C sütununda arama yap, C ve E sütunları - Tablo formatında"""
    try:
        if sheet_name not in workbook.sheetnames:
            return None, f"❌ '{sheet_name}' sayfası bulunamadı!"
        
        sheet = workbook[sheet_name]
        headers = get_column_headers(workbook, sheet_name)
        results = []
        search_lower = str(search_value).lower().strip()
        
        c_header = headers.get(3, "Ekipman Kodu")
        e_header = headers.get(5, "Kalibrasyon Tarihi")
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
            if len(row) > 2 and row[2] is not None:
                cell_value = str(row[2]).strip()
                if search_lower in cell_value.lower():
                    c_value = row[2] if len(row) > 2 else None
                    e_value = row[4] if len(row) > 4 else None
                    
                    formatted_date = format_date_to_tr(e_value) if e_value else "Belirtilmemiş"
                    
                    table = "```\n"
                    table += "┌─────────────────────┬─────────────────────────────────┐\n"
                    table += "│ Bilgi               │ Değer                           │\n"
                    table += "├─────────────────────┼─────────────────────────────────┤\n"
                    table += f"│ Satır No            │ {row_idx:<31} │\n"
                    table += f"│ {c_header:<17} │ {str(c_value)[:31]:<31} │\n"
                    table += f"│ {e_header:<17} │ {formatted_date:<31} │\n"
                    table += "└─────────────────────┴─────────────────────────────────┘\n"
                    table += "```"
                    results.append(table)
                    
                    if len(results) >= 10:
                        break
        
        if not results:
            return None, None
        return results, None
        
    except Exception as e:
        return None, f"Arama hatası: {str(e)}"

def search_by_date_range(start_date_tr, end_date_tr, workbook, sheet_name):
    """Belirli tarih aralığında kalibrasyonu yapılan ekipmanları listele"""
    try:
        if sheet_name not in workbook.sheetnames:
            return None, f"❌ '{sheet_name}' sayfası bulunamadı!"
        
        sheet = workbook[sheet_name]
        headers = get_column_headers(workbook, sheet_name)
        
        start_date = parse_date_tr(start_date_tr)
        end_date = parse_date_tr(end_date_tr)
        
        if not start_date or not end_date:
            return None, "❌ Hatalı tarih formatı! **gg.aa.yyyy** kullanın.\nÖrnek: 01.01.2026"
        
        c_header = headers.get(3, "Ekipman Kodu")
        e_header = headers.get(5, "Kalibrasyon Tarihi")
        
        found_equipment = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
            if len(row) > 4 and row[4] is not None:
                date_value = row[4]
                equipment_code = row[2] if len(row) > 2 else None
                description = row[3] if len(row) > 3 else None
                
                row_date = None
                if isinstance(date_value, str):
                    if '-' in date_value:
                        try:
                            row_date = datetime.strptime(date_value, "%Y-%m-%d")
                        except:
                            pass
                elif hasattr(date_value, 'strftime'):
                    row_date = date_value
                
                if row_date and start_date <= row_date <= end_date:
                    formatted_date = format_date_to_tr(date_value)
                    found_equipment.append({
                        'row': row_idx,
                        'code': equipment_code,
                        'description': description,
                        'date': formatted_date
                    })
        
        if not found_equipment:
            return None, f"❌ {start_date_tr} - {end_date_tr} tarihleri arasında kalibrasyon yapılan ekipman bulunamadı."
        
        # Sayfalama ile göster (her sayfada 20 satır)
        page_size = 20
        pages = []
        total_pages = (len(found_equipment) + page_size - 1) // page_size
        
        for page_num in range(total_pages):
            start_idx = page_num * page_size
            end_idx = min(start_idx + page_size, len(found_equipment))
            page_items = found_equipment[start_idx:end_idx]
            
            table = "```\n"
            table += f"┌────────┬─────────────────────────────────┬──────────────────────────────┬─────────────────┐\n"
            table += f"│ Satır  │ Ekipman Kodu                    │ Açıklama                     │ Kalibrasyon     │\n"
            table += f"│ No     │                                 │                              │ Tarihi          │\n"
            table += f"├────────┼─────────────────────────────────┼──────────────────────────────┼─────────────────┤\n"
            
            for item in page_items:
                code = str(item['code'])[:31] if item['code'] else "Belirtilmemiş"
                desc = str(item['description'])[:28] if item['description'] else "Belirtilmemiş"
                table += f"│ {item['row']:<6} │ {code:<31} │ {desc:<28} │ {item['date']:<15} │\n"
            
            table += f"└────────┴─────────────────────────────────┴──────────────────────────────┴─────────────────┘\n"
            table += "```"
            pages.append(table)
        
        return pages, f"✅ {len(found_equipment)} ekipman bulundu ({start_date_tr} - {end_date_tr}) ({total_pages} sayfa)"
        
    except Exception as e:
        return None, f"Arama hatası: {str(e)}"

def get_all_calibrated_equipment(workbook, sheet_name):
    """Tarihi olan tüm ekipmanları listele (C ve D sütunları) - sayfalama ile"""
    try:
        if sheet_name not in workbook.sheetnames:
            return None, f"❌ '{sheet_name}' sayfası bulunamadı!"
        
        sheet = workbook[sheet_name]
        headers = get_column_headers(workbook, sheet_name)
        
        c_header = headers.get(3, "Ekipman Kodu")
        d_header = headers.get(4, "Açıklama")
        
        equipment_list = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
            if len(row) > 4 and row[4] is not None:
                date_value = row[4]
                equipment_code = row[2] if len(row) > 2 else None
                description = row[3] if len(row) > 3 else None
                
                formatted_date = format_date_to_tr(date_value) if date_value else None
                
                if formatted_date:
                    equipment_list.append({
                        'row': row_idx,
                        'code': equipment_code,
                        'description': description,
                        'date': formatted_date
                    })
        
        if not equipment_list:
            return None, "❌ Henüz kalibrasyon tarihi eklenmiş ekipman bulunamadı."
        
        # Sayfalama ile göster (her sayfada 20 satır)
        page_size = 20
        pages = []
        total_pages = (len(equipment_list) + page_size - 1) // page_size
        
        for page_num in range(total_pages):
            start_idx = page_num * page_size
            end_idx = min(start_idx + page_size, len(equipment_list))
            page_items = equipment_list[start_idx:end_idx]
            
            table = "```\n"
            table += f"┌────────┬─────────────────────────────────┬──────────────────────────────┬─────────────────┐\n"
            table += f"│ Satır  │ Ekipman Kodu                    │ Açıklama                      │ Kalibrasyon     │\n"
            table += f"│ No     │                                 │                               │ Tarihi          │\n"
            table += f"├────────┼─────────────────────────────────┼──────────────────────────────┼─────────────────┤\n"
            
            for item in page_items:
                code = str(item['code'])[:31] if item['code'] else "Belirtilmemiş"
                desc = str(item['description'])[:28] if item['description'] else "Belirtilmemiş"
                table += f"│ {item['row']:<6} │ {code:<31} │ {desc:<28} │ {item['date']:<15} │\n"
            
            table += f"└────────┴─────────────────────────────────┴──────────────────────────────┴─────────────────┘\n"
            table += "```"
            pages.append(table)
        
        return pages, f"✅ Toplam {len(equipment_list)} ekipmanın kalibrasyon tarihi girilmiştir. ({total_pages} sayfa)"
        
    except Exception as e:
        return None, f"Hata: {str(e)}"

def update_calibration_date(equipment_code, new_date, workbook, sheet_name):
    """Excel'de kalibrasyon tarihini güncelle (E sütunu - 5. sütun)"""
    try:
        if sheet_name not in workbook.sheetnames:
            return False, f"Sayfa bulunamadı: {sheet_name}"
        
        sheet = workbook[sheet_name]
        search_lower = str(equipment_code).lower().strip()
        found = False
        updated_row = None
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, values_only=False), 3):
            if len(row) > 2 and row[2].value is not None:
                cell_value = str(row[2].value).strip()
                if cell_value.lower() == search_lower:
                    # E sütununu güncelle (5. sütun)
                    date_cell = sheet.cell(row_idx, 5)
                    date_cell.value = new_date
                    
                    # Yeşil renklendir
                    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                    date_cell.fill = green_fill
                    
                    found = True
                    updated_row = row_idx
                    break
        
        if not found:
            return False, f"Ekipman kodu bulunamadı: {equipment_code}"
        
        return True, f"Kalibrasyon tarihi güncellendi (Satır: {updated_row})"
        
    except Exception as e:
        return False, f"Hata: {str(e)}"

def clear_calibration_date(equipment_code, workbook, sheet_name):
    """Excel'de kalibrasyon tarihini temizle (E sütunu - 5. sütun)"""
    try:
        if sheet_name not in workbook.sheetnames:
            return False, f"Sayfa bulunamadı: {sheet_name}"
        
        sheet = workbook[sheet_name]
        search_lower = str(equipment_code).lower().strip()
        found = False
        updated_row = None
        old_value = None
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, values_only=False), 3):
            if len(row) > 2 and row[2].value is not None:
                cell_value = str(row[2].value).strip()
                if cell_value.lower() == search_lower:
                    # E sütununu temizle (5. sütun)
                    date_cell = sheet.cell(row_idx, 5)
                    old_value = date_cell.value
                    date_cell.value = None
                    
                    # Kırmızı renklendir (silindi)
                    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    date_cell.fill = red_fill
                    
                    logger.info(f"Tarih silindi: Satır {row_idx}, Eski değer: {old_value}")
                    
                    found = True
                    updated_row = row_idx
                    break
        
        if not found:
            return False, f"Ekipman kodu bulunamadı: {equipment_code}"
        
        return True, f"Kalibrasyon tarihi silindi (Satır: {updated_row})"
        
    except Exception as e:
        return False, f"Hata: {str(e)}"

def save_to_github(workbook):
    """Excel dosyasını GitHub'a kaydet"""
    try:
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        file_content = output.read()
        
        encoded_content = base64.b64encode(file_content).decode('utf-8')
        
        api_url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE_PATH}"
        
        headers = {
            "Authorization": f"token {GITHUB_PAT_TOKEN}",
            "Accept": "application/vnd.github.v3+json"
        }
        
        get_response = requests.get(api_url, headers=headers)
        sha = None
        if get_response.status_code == 200:
            sha = get_response.json().get("sha")
        
        commit_data = {
            "message": f"Bot ile kalibrasyon tarihi güncellendi - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "content": encoded_content,
            "branch": "main"
        }
        if sha:
            commit_data["sha"] = sha
        
        put_response = requests.put(api_url, headers=headers, json=commit_data)
        
        if put_response.status_code in [200, 201]:
            return True, "GitHub'a kaydedildi"
        else:
            return False, f"Kayıt hatası: {put_response.status_code}"
            
    except Exception as e:
        return False, f"Hata: {str(e)}"

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("🔍 Excel bot başlatılıyor...\n\nExcel dosyası kontrol ediliyor...")
    
    try:
        response = requests.get(EXCEL_URL, timeout=30)
        workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        
        sheets = workbook.sheetnames
        sheet_found = SHEET_NAME in sheets
        
        info_text = f"📊 *Excel Bilgileri*\n\n"
        info_text += f"• Sayfalar: {', '.join(sheets)}\n"
        info_text += f"• Aranan sayfa '{SHEET_NAME}': {'✅ Bulundu' if sheet_found else '❌ Bulunamadı'}\n"
        
        if sheet_found:
            sheet = workbook[SHEET_NAME]
            info_text += f"• Toplam satır: {sheet.max_row}\n"
            info_text += f"• Toplam sütun: {sheet.max_column}\n\n"
            
            headers = get_column_headers(workbook, SHEET_NAME)
            info_text += "📝 *Kolon Başlıkları (2. satır)*\n"
            for col_idx, header in list(headers.items())[:8]:
                info_text += f"• {header}\n"
        
        if GITHUB_PAT_TOKEN:
            info_text += f"\n✅ GitHub entegrasyonu aktif"
        else:
            info_text += f"\n⚠️ GitHub entegrasyonu pasif"
        
        await update.message.reply_text(info_text, parse_mode="Markdown")
        
    except Exception as e:
        await update.message.reply_text(f"⚠️ Excel okuma hatası: {str(e)}")

async def search_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("❌ *Kullanım:* `/ara ARANACAK_DEGER`\n\nÖrnek: `/ara 00GHC`\n\nBu komut kısmi eşleşme yapar ve sonuçları tablo olarak gösterir.", parse_mode="Markdown")
        return
    
    search_text = " ".join(context.args).strip()
    
    await update.message.reply_text(f"🔍 '{search_text}' aranıyor... (C sütununda, **kısmi eşleşme**)", parse_mode="Markdown")
    
    try:
        response = requests.get(EXCEL_URL, timeout=30)
        response.raise_for_status()
        
        workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        
        results, error = search_in_column_c_partial(search_text, workbook, SHEET_NAME)
        
        if error:
            await update.message.reply_text(error)
        elif results:
            await update.message.reply_text(f"✅ {len(results)} sonuç bulundu:\n\n" + "\n".join(results), parse_mode="Markdown")
        else:
            await update.message.reply_text(f"❌ '{search_text}' için C sütununda eşleşme bulunamadı.")
            
    except Exception as e:
        await update.message.reply_text(f"❌ Hata: {str(e)}")

async def tarih_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("❌ *Kullanım:* `/tarih ARANACAK_DEGER`\n\nÖrnek: `/tarih 00GHC01CP101`\n\nBu komut sadece ekipman kodu (C) ve kalibrasyon tarihini (E) gösterir.", parse_mode="Markdown")
        return
    
    search_text = " ".join(context.args).strip()
    
    await update.message.reply_text(f"📅 '{search_text}' için kalibrasyon tarihi aranıyor...")
    
    try:
        response = requests.get(EXCEL_URL, timeout=30)
        workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        
        results, error = search_calibration_date(search_text, workbook, SHEET_NAME)
        
        if error:
            await update.message.reply_text(error)
        elif results:
            await update.message.reply_text(f"✅ *Kalibrasyon Bilgileri*\n\n" + "\n".join(results), parse_mode="Markdown")
        else:
            await update.message.reply_text(f"❌ '{search_text}' için kayıt bulunamadı.")
            
    except Exception as e:
        await update.message.reply_text(f"❌ Hata: {str(e)}")

async def tarih_ara_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Belirli tarih aralığında kalibrasyonu yapılan ekipmanları listele"""
    if len(context.args) < 2:
        await update.message.reply_text("❌ *Kullanım:* `/tarih_ara BASLANGIC_TARIH BITIS_TARIH`\n\nÖrnek: `/tarih_ara 01.01.2026 31.12.2026`\n\nTarih formatı: **gg.aa.yyyy**", parse_mode="Markdown")
        return
    
    start_date = context.args[0].strip()
    end_date = context.args[1].strip()
    
    if not re.match(r'^\d{2}\.\d{2}\.\d{4}$', start_date) or not re.match(r'^\d{2}\.\d{2}\.\d{4}$', end_date):
        await update.message.reply_text("❌ Hatalı tarih formatı! Lütfen **gg.aa.yyyy** formatında girin.\nÖrnek: 01.01.2026", parse_mode="Markdown")
        return
    
    await update.message.reply_text(f"📅 '{start_date}' - '{end_date}' tarihleri arasında kalibrasyon yapılan ekipmanlar aranıyor...")
    
    try:
        response = requests.get(EXCEL_URL, timeout=30)
        workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        
        results, message = search_by_date_range(start_date, end_date, workbook, SHEET_NAME)
        
        if results:
            await update.message.reply_text(message, parse_mode="Markdown")
            for result in results:
                await update.message.reply_text(result, parse_mode="Markdown")
                await asyncio.sleep(0.3)
        else:
            await update.message.reply_text(message)
            
    except Exception as e:
        await update.message.reply_text(f"❌ Hata: {str(e)}")

async def listeli_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Tarihi olan tüm ekipmanları listele (C ve D sütunları) - sayfalama ile"""
    await update.message.reply_text("📋 Kalibrasyon tarihi girilmiş tüm ekipmanlar listeleniyor...\n\nBu işlem birkaç saniye sürebilir.")
    
    try:
        response = requests.get(EXCEL_URL, timeout=30)
        workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        
        results, message = get_all_calibrated_equipment(workbook, SHEET_NAME)
        
        if results:
            await update.message.reply_text(message, parse_mode="Markdown")
            for result in results:
                await update.message.reply_text(result, parse_mode="Markdown")
                await asyncio.sleep(0.3)
        else:
            await update.message.reply_text(message)
            
    except Exception as e:
        await update.message.reply_text(f"❌ Hata: {str(e)}")

async def guncelle_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kalibrasyon tarihini güncelle (gg.aa.yyyy formatında)"""
    if len(context.args) < 2:
        await update.message.reply_text("❌ *Kullanım:* `/guncelle EKIPMAN_KODU YENI_TARIH`\n\nÖrnek: `/guncelle 00GHC01CP101 14.04.2026`\n\nTarih formatı: **gg.aa.yyyy**", parse_mode="Markdown")
        return
    
    equipment_code = context.args[0].strip()
    new_date_tr = context.args[1].strip()
    
    if not re.match(r'^\d{2}\.\d{2}\.\d{4}$', new_date_tr):
        await update.message.reply_text("❌ Hatalı tarih formatı! Lütfen **gg.aa.yyyy** formatında girin.\nÖrnek: 14.04.2026", parse_mode="Markdown")
        return
    
    converted_date, is_valid = validate_date_tr(new_date_tr)
    if not is_valid:
        await update.message.reply_text("❌ Geçersiz tarih! Lütfen gerçek bir tarih girin.\nÖrnek: 14.04.2026", parse_mode="Markdown")
        return
    
    await update.message.reply_text(f"✏️ '{equipment_code}' için kalibrasyon tarihi güncelleniyor: {new_date_tr}")
    
    try:
        response = requests.get(EXCEL_URL, timeout=30)
        workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=False)
        
        success, message = update_calibration_date(equipment_code, converted_date, workbook, SHEET_NAME)
        
        if not success:
            await update.message.reply_text(f"❌ {message}")
            return
        
        if GITHUB_PAT_TOKEN:
            github_success, github_message = save_to_github(workbook)
            if github_success:
                await update.message.reply_text(f"✅ {message}\n\n📤 **{github_message}**\n\n• Ekipman: `{equipment_code}`\n• Yeni Tarih: `{new_date_tr}`", parse_mode="Markdown")
            else:
                await update.message.reply_text(f"✅ {message}\n\n❌ GitHub hatası: {github_message}", parse_mode="Markdown")
        else:
            await update.message.reply_text(f"✅ {message}\n\n⚠️ GitHub token bulunamadı! Değişiklikler KAYDEDİLMEDİ.", parse_mode="Markdown")
            
    except Exception as e:
        await update.message.reply_text(f"❌ Hata: {str(e)}")

async def sil_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kalibrasyon tarihini sil"""
    if not context.args:
        await update.message.reply_text("❌ *Kullanım:* `/sil EKIPMAN_KODU`\n\nÖrnek: `/sil 00GHC01CP101`\n\nBu komut kalibrasyon tarihini tamamen siler.", parse_mode="Markdown")
        return
    
    equipment_code = context.args[0].strip()
    
    await update.message.reply_text(f"🗑️ '{equipment_code}' için kalibrasyon tarihi siliniyor...")
    
    try:
        response = requests.get(EXCEL_URL, timeout=30)
        workbook = openpyxl.load_workbook(BytesIO(response.content), data_only=False)
        
        success, message = clear_calibration_date(equipment_code, workbook, SHEET_NAME)
        
        if not success:
            await update.message.reply_text(f"❌ {message}")
            return
        
        if GITHUB_PAT_TOKEN:
            github_success, github_message = save_to_github(workbook)
            if github_success:
                await update.message.reply_text(f"✅ {message}\n\n📤 **{github_message}**\n\n• Ekipman: `{equipment_code}`\n• Tarih silindi.", parse_mode="Markdown")
            else:
                await update.message.reply_text(f"✅ {message}\n\n❌ GitHub hatası: {github_message}", parse_mode="Markdown")
        else:
            await update.message.reply_text(f"✅ {message}\n\n⚠️ GitHub token bulunamadı! Değişiklikler KAYDEDİLMEDİ.", parse_mode="Markdown")
            
    except Exception as e:
        await update.message.reply_text(f"❌ Hata: {str(e)}")

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    help_text = """📖 *Yardım Menüsü*

*Komutlar:*

🔍 `/ara <deger>` - **C sütununda kısmi eşleşme** arama yapar
   • Tüm sütunları **tablo** olarak gösterir
   • Örnek: `/ara 00GHC`

📅 `/tarih <deger>` - Tek ekipman için kalibrasyon tarihi sorgular
   • Sadece **C (ekipman kodu)** ve **E (kalibrasyon tarihi)** gösterir
   • Örnek: `/tarih 00GHC01CP101`

📆 `/tarih_ara <baslangic> <bitis>` - Tarih aralığında kalibrasyon yapılan ekipmanları listeler
   • Örnek: `/tarih_ara 01.01.2026 31.12.2026`

📋 `/listeli` - **Kalibrasyon tarihi girilmiş TÜM ekipmanları** listeler
   • C (ekipman kodu) ve D (açıklama) sütunlarını gösterir
   • Toplam sayıyı da belirtir
   • Uzun listelerde **sayfalama** yapar

✏️ `/guncelle <kod> <tarih>` - Kalibrasyon tarihini günceller
   • Örnek: `/guncelle 00GHC01CP101 14.04.2026`

🗑️ `/sil <kod>` - Kalibrasyon tarihini **tamamen siler**
   • Örnek: `/sil 00GHC01CP101`

ℹ️ `/start` - Botu başlat ve Excel bilgilerini göster
🆘 `/help` - Bu yardım menüsü

*Tarih Formatı:* **gg.aa.yyyy** (örnek: 14.04.2026)
*Not:* Tüm tarih girişleri ve çıkışları bu formattadır!"""
    
    await update.message.reply_text(help_text, parse_mode="Markdown")

def main():
    if not TOKEN:
        logger.error("❌ TELEGRAM_BOT_TOKEN yok!")
        return
    
    if not EXCEL_URL:
        logger.error("❌ EXCEL_URL yok!")
        return
    
    logger.info(f"Excel URL: {EXCEL_URL}")
    logger.info(f"Aranan sayfa: {SHEET_NAME}")
    logger.info(f"GitHub Repo: {GITHUB_REPO}")
    logger.info(f"GitHub Token: {'✅ Var' if GITHUB_PAT_TOKEN else '❌ Yok'}")
    
    application = Application.builder().token(TOKEN).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("ara", search_command))
    application.add_handler(CommandHandler("tarih", tarih_command))
    application.add_handler(CommandHandler("tarih_ara", tarih_ara_command))
    application.add_handler(CommandHandler("listeli", listeli_command))
    application.add_handler(CommandHandler("guncelle", guncelle_command))
    application.add_handler(CommandHandler("sil", sil_command))
    application.add_handler(CommandHandler("help", help_command))
    
    application.run_polling()

if __name__ == "__main__":
    main()
